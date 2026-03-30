import requests
import json
import csv
from io import BytesIO, StringIO
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import load_workbook
from rich import print

class MockUser:
    is_authenticated = False
    is_staff = False

class MockFiles(dict):
    def getlist(self, key):
        val = self.get(key, [])
        if not isinstance(val, list):
            return [val]
        return val

class MockRequest:
    def __init__(self, data=None, files=None):
        self.data = data or {}
        self.FILES = files or MockFiles()
        self.user = MockUser()

    def build_absolute_uri(self, location):
        return location

def download_image_as_file(url, field_name="image"):
    try:
        url = url.strip()
        if not url.startswith("http"):
            return None
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            filename = url.split("/")[-1]
            if not filename or "." not in filename:
                filename = f"{field_name}.jpg"
            return SimpleUploadedFile(
                name=filename,
                content=response.content,
                content_type=response.headers.get("Content-Type", "image/jpeg")
            )
    except Exception as e:
        print(f"Error downloading image {url}: {e}")
    return None

def parse_file_to_dicts(file_obj, file_extension):
    data_rows = []
    
    if file_extension == 'json':
        try:
            raw_data = json.load(file_obj)
            if isinstance(raw_data, dict):
                raw_data = [raw_data]
            for item in raw_data:
                lowered_item = {str(k).lower().strip(): v for k, v in item.items()}
                data_rows.append(lowered_item)
        except Exception as e:
            print(f"Error parsing JSON: {e}")
            
    elif file_extension == 'csv':
        try:
            # decode using utf-8-sig to handle BOM added by Excel
            decoded_file = file_obj.read().decode('utf-8-sig').splitlines()
            reader = csv.reader(decoded_file)
            rows = list(reader)
            if rows:
                headers = [str(h).lower().strip() for h in rows[0]]
                for row in rows[1:]:
                    if all(c is None or str(c).strip() == '' for c in row):
                        continue
                    data_rows.append(dict(zip(headers, row)))
        except Exception as e:
            print(f"Error parsing CSV: {e}")
            
    elif file_extension in ['xls', 'xlsx']:
        try:
            wb = load_workbook(file_obj)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h).lower().strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
                for row in rows[1:]:
                    if all(c is None or str(c).strip() == '' for c in row):
                        continue
                    data_rows.append(dict(zip(headers, row)))
        except Exception as e:
            print(f"Error parsing Excel: {e}")
            
    return data_rows

def process_import_file(file_obj, file_extension, serializer_class):
    rows_as_dicts = parse_file_to_dicts(file_obj, file_extension)
    
    if not rows_as_dicts:
        return 0, ["Empty file or invalid format"]

    created = 0
    errors = []

    image_fields = ['image', 'logo', 'uploaded_images', 'gallery_images']

    for data in rows_as_dicts:
        print(data)
        
        files = MockFiles()
        
        # Process potential image URLs
        for field in image_fields:
            if field in data and data[field]:
                val = str(data[field]).strip()
                
                validated_files = None
                if val.startswith("http"):
                    # split by comma for multiple images
                    urls = [u for u in val.split(',') if u.strip().startswith('http')]
                    downloaded_files = []
                    for u in urls:
                        f = download_image_as_file(u, field)
                        if f:
                            downloaded_files.append(f)
                    
                    if downloaded_files:
                        if len(downloaded_files) == 1 and field != 'uploaded_images':
                            # Single file field
                            validated_files = downloaded_files[0]
                        else:
                            # List of files
                            validated_files = downloaded_files
                
                if validated_files:
                    files[field] = validated_files
                    data[field] = validated_files
                else:
                    # Clear invalid values so DRF doesn't treat strings as invalid files
                    data.pop(field, None)

        mock_req = MockRequest(data=data, files=files)

        serializer = serializer_class(data=data, context={'request': mock_req})
        
        if serializer.is_valid():
            serializer.save()
            created += 1
        else:
            errors.append(serializer.errors)

    return created, errors